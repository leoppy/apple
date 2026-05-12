import json
import os
import time
import requests
from datetime import datetime
from dotenv import load_dotenv
from pathlib import Path
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


@dataclass
class JiraConfig:
    """Jira 配置类"""
    token: str
    node_id: int  # 必须指定的节点ID
    project_id: int  # 项目ID，API 请求必需参数
    base_url: str = "https://jira.i-soft.com.cn"
    api_endpoint: str = "/rest/com.easesolutions.jira.plugins.requirements/1.0/tree/data"
    delay_ms: int = 500  # 请求延迟（毫秒）
    output_dir: Path = Path(__file__).parent / "tempFile"

    @classmethod
    def from_env(cls, node_id: int, project_id: Optional[int] = None, delay_ms: Optional[int] = None):
        """从环境变量加载配置"""
        env_path = Path(__file__).parent.parent.parent / "02_skills" / "token-manager" / ".env"
        load_dotenv(dotenv_path=env_path)

        token = os.getenv("JIRA_TOKEN")
        if not token:
            raise ValueError("请在 .env 文件中配置 JIRA_TOKEN")

        # 从环境变量读取默认值（如果未通过参数指定）
        if project_id is None:
            project_id = int(os.getenv("JIRA_PROJECT_ID", "10002"))

        if delay_ms is None:
            delay_ms = int(os.getenv("JIRA_DELAY_MS", "500"))

        return cls(token=token, node_id=node_id, project_id=project_id, delay_ms=delay_ms)

    @property
    def full_api_url(self):
        return f"{self.base_url}{self.api_endpoint}"

    @property
    def delay_seconds(self):
        """将毫秒转换为秒"""
        return self.delay_ms / 1000


class JiraRequirementsExporter:
    """Jira Requirements 导出器"""

    def __init__(self, config: JiraConfig):
        self.config = config
        self.items_data: List[Dict] = []
        self.req_list: List[Tuple[int, int]] = [(config.node_id, 0)]
        self.processed_count = 0

    def _make_request(self, item_id: int) -> Dict:
        """发起 API 请求"""
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.config.token}",
            "X-Atlassian-Token": "no-check",
        }

        data = {
            'folderId': item_id,
            "offset": 0,
            "projectId": self.config.project_id,
            "queryParams": f"{item_id}"
        }

        response = requests.post(
            self.config.full_api_url,
            data=json.dumps(data),
            headers=headers,
            timeout=30
        )

        if response.status_code != 200:
            raise Exception(f"API 请求失败，状态码: {response.status_code}, 响应: {response.text[:200]}")

        return response.json()

    def _process_item(self, item: Dict, parent_id: int, level: int):
        """处理单个项目"""
        # 判断类型：有 key 是 Issue，否则是 Folder
        item_type = 'Issue' if item.get('key') else 'Folder'

        item_data = {
            'id': item.get('id'),
            'key': item.get('key', ''),
            'name': item.get('name', ''),
            'type': item_type,
            'parent_id': parent_id,
            'level': level,
            'position': item.get('position', 0),
            'description': item.get('description', '').strip()
        }

        self.items_data.append(item_data)

        # 如果有子节点，加入待处理队列
        if item.get('hasChild'):
            self.req_list.append((int(item['id']), level + 1))

    def export(self):
        """执行导出"""
        start_time = time.time()
        print(f'开始导出 Jira Requirements 数据 [{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}]')
        print(f'节点ID: {self.config.node_id}')
        print(f'API: {self.config.full_api_url}')
        print('-' * 60)

        # 初始化队列：(item_id, level)
        self.req_list = [(self.config.node_id, 0)]

        while self.req_list:
            item_id, level = self.req_list.pop(0)
            self.processed_count += 1

            print(f'[{self.processed_count}] 正在处理节点 {item_id} (层级 {level})...')

            try:
                resp_data = self._make_request(item_id)
                item_list = resp_data.get("itemList", [])

                for item in item_list:
                    self._process_item(item, item_id, level + 1)

                time.sleep(self.config.delay_seconds)

            except Exception as e:
                print(f'  ❌ 错误: {e}')
                continue

        end_time = time.time()
        duration = int(end_time - start_time)

        print('-' * 60)
        print(f'✅ 导出完成!')
        print(f'   总用时: {duration} 秒')
        print(f'   处理节点数: {self.processed_count}')
        print(f'   收集数据条数: {len(self.items_data)}')

        return self.items_data

    def save_to_files(self):
        """保存到 Excel 文件（带格式）"""
        # 确保输出目录存在
        self.config.output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"jira_requirements_{self.config.node_id}_{timestamp}.xlsx"
        excel_path = self.config.output_dir / excel_filename

        # 创建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "Jira Requirements"

        if not self.items_data:
            print("   ⚠️  没有数据可导出")
            return

        # 写入表头
        headers = list(self.items_data[0].keys())
        ws.append(headers)

        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 写入数据
        for item in self.items_data:
            ws.append(list(item.values()))

        # 设置列宽
        column_widths = {
            'id': 10,
            'key': 15,
            'name': 40,
            'type': 12,
            'parent_id': 12,
            'level': 8,
            'position': 10,
            'description': 75  # 约 20cm (1cm ≈ 3.75 字符宽度)
        }

        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = column_widths.get(header, 15)

        # 冻结首行
        ws.freeze_panes = "A2"

        # 添加自动筛选
        ws.auto_filter.ref = ws.dimensions

        # 保存文件
        wb.save(excel_path)
        print(f'   Excel 文件已保存: {excel_path}')


def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(
        description='导出 Jira Requirements 数据到 Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 基本用法（使用默认项目ID 10002）
  python main.py 11693

  # 指定项目ID
  python main.py 11693 --project-id 10003

  # 指定延迟（毫秒）
  python main.py 11693 --delay 1000

  # 完整参数
  python main.py 11693 --project-id 10002 --delay 500
        """
    )

    parser.add_argument('node_id', type=int, help='节点ID（必需）')
    parser.add_argument('--project-id', '-p', type=int, help='项目ID（默认从环境变量读取，或使用 10002）')
    parser.add_argument('--delay', '-d', type=int, help='请求延迟（毫秒，默认 500）')

    args = parser.parse_args()

    config = JiraConfig.from_env(
        node_id=args.node_id,
        project_id=args.project_id,
        delay_ms=args.delay
    )

    print(f'配置信息:')
    print(f'  节点ID: {config.node_id}')
    print(f'  项目ID: {config.project_id}')
    print(f'  延迟: {config.delay_ms}ms')
    print()

    exporter = JiraRequirementsExporter(config)
    exporter.export()
    exporter.save_to_files()


if __name__ == '__main__':
    main()
