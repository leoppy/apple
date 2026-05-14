from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Requirement, TestCase, TraceIssue


class ReportGenerator:
    def __init__(
        self,
        output_cfg: Dict,
        project_root: Path,
        testcases: List[TestCase],
        requirements: Dict[str, Requirement],
        issues: List[TraceIssue],
        coverage_detail_rows: List[Dict],
        coverage_summary_rows: List[Dict],
        orphan_rows: List[Dict],
        pass_rate_rows: List[Dict],
        module_overview_rows: List[Dict],
    ):
        self.output_cfg = output_cfg
        self.project_root = project_root.resolve()
        self.testcases = testcases
        self.requirements = requirements
        self.issues = issues
        self.coverage_detail_rows = coverage_detail_rows
        self.coverage_summary_rows = coverage_summary_rows
        self.orphan_rows = orphan_rows
        self.pass_rate_rows = pass_rate_rows
        self.module_overview_rows = module_overview_rows

    def generate_report(
        self,
        output_filename: str | None = None,
        report_type: str = "coverage",
    ) -> Path:
        date_dir = datetime.now().strftime("%Y%m%d")
        output_dir = self.project_root / "tempFile" / "reports" / date_dir
        output_dir.mkdir(parents=True, exist_ok=True)

        if output_filename:
            filename = output_filename
        else:
            # 从需求中提取唯一的 level 列表
            levels = sorted(set(req.level for req in self.requirements.values()))
            level_str = "_".join(levels) if levels else "unknown"

            # 根据报告类型生成文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if report_type == "coverage":
                filename = f"report_{level_str}_coverage_{timestamp}.xlsx"
            elif report_type == "pass-rate":
                filename = f"report_{level_str}_passRate_{timestamp}.xlsx"
            else:
                filename = f"report_{level_str}_{timestamp}.xlsx"

        output_path = output_dir / filename

        wb = Workbook()
        wb.remove(wb.active)

        if report_type == "coverage":
            self.create_module_overview_sheet(wb)
            self.create_coverage_summary_sheet(wb)
            self.create_coverage_detail_sheet(wb)
            self.create_orphan_traces_sheet(wb)
        elif report_type == "pass-rate":
            self.create_pass_rate_sheet(wb)

        wb.save(output_path)
        return output_path

        wb.save(output_path)
        return output_path

    def create_trace_matrix_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("追溯矩阵")
        headers = ["用例ID", "用例名称", "测试类型", "追溯需求", "测试结果", "来源文件", "Confluence页面"]
        ws.append(headers)

        orphan_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        mismatch_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        for tc in self.testcases:
            keys = sorted(tc.traced_keys) if tc.traced_keys else ["-"]
            for key in keys:
                row = [tc.case_id, tc.case_name, tc.test_type, key, tc.result, tc.source_file, tc.source_page]
                ws.append(row)
                row_num = ws.max_row
                issue_list = self.issue_index.get((tc.case_id, key), [])
                issue_types = {issue.issue_type for issue in issue_list}
                if "orphan" in issue_types:
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=row_num, column=c).fill = orphan_fill
                elif "mismatch" in issue_types:
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=row_num, column=c).fill = mismatch_fill

        self._format_sheet(ws)

    def create_module_overview_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("模块概览")
        headers = ["模块", "未覆盖需求数", "未覆盖需求列表"]
        ws.append(headers)
        for row in self.module_overview_rows:
            ws.append(
                [
                    row["module"],
                    row["uncovered_count"],
                    row["uncovered_keys"],
                ]
            )
        self._format_sheet(ws)

    def create_coverage_summary_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("覆盖率汇总")
        headers = ["项目", "层级", "总需求数", "已覆盖", "未覆盖", "覆盖率"]
        ws.append(headers)

        for row in self.coverage_summary_rows:
            ws.append(
                [
                    row["project"],
                    row["level"],
                    row["total_requirements"],
                    row["covered_requirements"],
                    row["uncovered_requirements"],
                    row["coverage_rate"],
                ]
            )

        self._format_sheet(ws, percentage_columns={6})

    def create_coverage_detail_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("覆盖率明细")
        headers = ["需求Key", "模块", "类型", "标签", "优先级", "是否覆盖"]
        ws.append(headers)

        for row in self.coverage_detail_rows:
            ws.append(
                [
                    row["requirement_key"],
                    row["module"],
                    row["issue_type"],
                    row["labels"],
                    row["priority"],
                    row["is_covered"],
                ]
            )

        self._format_sheet(ws)

    def create_issues_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("问题清单")
        headers = ["严重性", "问题类型", "用例ID", "需求Key", "模块", "问题描述"]
        ws.append(headers)
        severity_order = {"high": 0, "medium": 1, "low": 2}
        sorted_issues = sorted(self.issues, key=lambda x: severity_order.get(x.severity, 99))
        for issue in sorted_issues:
            ws.append([issue.severity, issue.issue_type, issue.testcase_id, issue.requirement_key, issue.components, issue.description])
        self._format_sheet(ws)

    def create_pass_rate_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("通过率统计")
        headers = ["模块", "总用例数", "通过", "失败", "未执行", "通过率"]
        ws.append(headers)
        for row in self.pass_rate_rows:
            ws.append(
                [
                    row["module"],
                    row["total"],
                    row["passed"],
                    row["failed"],
                    row["not_executed"],
                    row["pass_rate"],
                ]
            )
        self._format_sheet(ws, percentage_columns={6})

    def create_orphan_traces_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("孤立追溯")
        headers = ["需求Key", "追溯用例数", "用例ID列表", "模块"]
        ws.append(headers)
        for row in self.orphan_rows:
            ws.append(
                [
                    row["requirement_key"],
                    row["testcase_count"],
                    row["testcase_ids"],
                    row["modules"],
                ]
            )
        self._format_sheet(ws)

    @staticmethod
    def _format_sheet(ws, percentage_columns: set[int] | None = None) -> None:
        percentage_columns = percentage_columns or set()
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                value = ws.cell(row=row, column=col).value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 10), 60)

        for row in range(2, ws.max_row + 1):
            for col in percentage_columns:
                cell = ws.cell(row=row, column=col)
                cell.number_format = "0.00%"
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
