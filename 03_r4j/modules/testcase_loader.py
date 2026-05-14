from __future__ import annotations

import json
import logging
import re
import time
from pathlib import Path
from typing import Dict, List, Set

from openpyxl import load_workbook

from .confluence_client import ConfluenceClient
from .models import TestCase

ISSUE_KEY_PATTERN = re.compile(r"[A-Z]+-\d+")


class TestCaseLoader:
    def __init__(self, config: Dict, client: ConfluenceClient, use_cache: bool = True):
        self.config = config
        self.client = client
        self.use_cache = use_cache
        self.logger = logging.getLogger(__name__)
        self.project_root = Path(config.get("_project_root", Path(__file__).resolve().parents[1])).resolve()
        cache_cfg = config.get("cache", {})
        self.cache_enabled = bool(cache_cfg.get("enabled", True)) and use_cache
        cache_base_dir = Path(cache_cfg.get("dir", "tempFile/cache"))
        if cache_base_dir.is_absolute():
            resolved_cache_base = cache_base_dir
        else:
            resolved_cache_base = (self.project_root / cache_base_dir).resolve()

        config_cache_namespace = str(cache_cfg.get("namespace", "")).strip()
        if not config_cache_namespace:
            config_cache_namespace = str(config.get("_config_name", "")).strip()
        if not config_cache_namespace:
            config_cache_namespace = "default"

        self.cache_dir = resolved_cache_base / config_cache_namespace
        self.cache_ttl_hours = int(cache_cfg.get("ttl_hours", 24))

    def load_all_testcases(self) -> List[TestCase]:
        all_cases: List[TestCase] = []
        for tc_cfg in self.config.get("confluence_testcases", []):
            all_cases.extend(self._load_from_page_config(tc_cfg))
        return all_cases

    def _load_from_page_config(self, tc_cfg: Dict) -> List[TestCase]:
        page_id = str(tc_cfg["page_id"])
        root_page = self.client.get_page(page_id)
        pages = [root_page]
        if tc_cfg.get("traverse_children", False):
            pages.extend(self.client.get_descendant_pages(page_id))

        cases: List[TestCase] = []
        for page in pages:
            current_page_id = str(page["id"])
            current_page_title = page.get("title", "")
            attachments = self.client.get_page_attachments(current_page_id)
            excel_attachments = [
                item
                for item in attachments
                if str(item.get("title", "")).lower().endswith((".xlsx", ".xlsm", ".xls"))
            ]

            for attachment in excel_attachments:
                excel_path = self._get_excel_file(current_page_id, attachment)
                cases.extend(
                    self.parse_excel(
                        excel_path=excel_path,
                        config=tc_cfg,
                        source_page=current_page_title,
                    )
                )
        return cases

    def _get_excel_file(self, page_id: str, attachment: Dict) -> Path:
        attachment_id = str(attachment.get("id"))
        attachment_version = str(attachment.get("version", {}).get("number", "0"))
        attachment_name = str(attachment.get("title", "attachment.xlsx"))
        safe_name = attachment_name.replace("\\", "_").replace("/", "_")
        cached_name = f"testcases_{page_id}_{attachment_id}_{attachment_version}_{safe_name}"
        cache_file = self.cache_dir / "testcases" / cached_name

        if self.cache_enabled and cache_file.exists():
            age_seconds = time.time() - cache_file.stat().st_mtime
            if age_seconds <= self.cache_ttl_hours * 3600:
                return cache_file

        downloaded = self.client.download_attachment(attachment=attachment, save_path=cache_file)
        return downloaded

    def parse_excel(self, excel_path: Path, config: Dict, source_page: str) -> List[TestCase]:
        wb = load_workbook(excel_path, data_only=True, read_only=True)
        sheet_pattern = re.compile(config.get("sheet_pattern", ".*测试用例.*"))
        excel_columns = config.get("excel_columns", {})
        result_mapping = config.get("result_mapping", {})
        test_type = str(config.get("test_type", "")).strip()
        module = self._extract_module_from_filename(excel_path.name)

        testcases: List[TestCase] = []
        selected_sheets = [name for name in wb.sheetnames if sheet_pattern.search(name)]
        if not selected_sheets:
            self.logger.warning("文件 %s 未命中 sheet_pattern，回退解析全部 Sheet", excel_path.name)
            selected_sheets = list(wb.sheetnames)

        for sheet_name in selected_sheets:
            if not sheet_pattern.search(sheet_name):
                self.logger.debug("Sheet %s 非 pattern 命中，走回退解析", sheet_name)
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header_idx, header_map = self._find_header_map(rows, excel_columns)
            if header_idx < 0:
                self.logger.warning("文件 %s / Sheet %s 未识别到表头，已跳过", excel_path.name, sheet_name)
                continue

            id_col = self._find_col(header_map, [excel_columns.get("id", ""), "用例ID", "测试用例ID", "用例编号", "ID"])
            name_col = self._find_col(header_map, [excel_columns.get("name", ""), "用例名称", "测试项名称", "测试用例名称"])
            trace_col = self._find_col(
                header_map,
                [excel_columns.get("trace", ""), "对应需求编号", "对应需求ID", "追溯需求", "关联需求", "需求编号", "需求ID"],
            )
            result_col = self._find_col(header_map, [excel_columns.get("result", ""), "测试结果", "执行结果", "结果"])
            if id_col is None or trace_col is None:
                self.logger.warning("文件 %s / Sheet %s 列映射失败(id/trace)", excel_path.name, sheet_name)
                continue

            for row in rows[header_idx + 1 :]:
                case_id = self._safe_cell(row, id_col)
                if not case_id:
                    continue
                case_name = self._safe_cell(row, name_col) if name_col is not None else ""
                trace_raw = self._safe_cell(row, trace_col)
                result_raw = self._safe_cell(row, result_col) if result_col is not None else ""
                traced_keys = self.extract_traced_keys(trace_raw)
                mapped_result = self._map_result(result_raw, result_mapping)
                testcases.append(
                    TestCase(
                        case_id=case_id,
                        case_name=case_name,
                        test_type=test_type,
                        traced_keys=traced_keys,
                        result=mapped_result,
                        source_file=excel_path.name,
                        source_page=source_page,
                        module=module,
                    )
                )
        wb.close()
        return testcases

    @staticmethod
    def _normalize_header(value: str) -> str:
        return re.sub(r"\s+", "", str(value or "")).replace("\u3000", "").strip().lower()

    @staticmethod
    def _extract_module_from_filename(filename: str) -> str:
        """\u4ece\u6587\u4ef6\u540d\u4e2d\u63d0\u53d6\u6a21\u5757\u540d
        \u683c\u5f0f: testcases_106154865_106155659_3_[PH-PRD-QC-007-2024]_\u8f6f\u4ef6\u6d4b\u8bd5\u89c4\u8303\u53ca\u62a5\u544a_BswM_TC4D9.xlsx
        \u63d0\u53d6"\u89c4\u8303\u53ca\u62a5\u544a"\u4e4b\u540e\u3001TC\u5f00\u5934\u4e4b\u524d\u7684\u90e8\u5206\u4f5c\u4e3a\u6a21\u5757\u540d
        """
        if not filename:
            return ""
        # \u53bb\u6389\u6269\u5c55\u540d
        name_without_ext = filename.rsplit(".", 1)[0] if "." in filename else filename
        # \u6309\u4e0b\u5212\u7ebf\u5206\u5272
        parts = name_without_ext.split("_")

        # \u627e\u5230"\u89c4\u8303\u53ca\u62a5\u544a"\u7684\u4f4d\u7f6e
        report_idx = -1
        for i, part in enumerate(parts):
            if "\u89c4\u8303\u53ca\u62a5\u544a" in part or "\u62a5\u544a" in part:
                report_idx = i
                break

        # \u4ece"\u89c4\u8303\u53ca\u62a5\u544a"\u4e4b\u540e\u5f00\u59cb\u67e5\u627e\u6a21\u5757\u540d
        if report_idx >= 0 and report_idx + 1 < len(parts):
            # \u63d0\u53d6"\u89c4\u8303\u53ca\u62a5\u544a"\u4e4b\u540e\u7684\u7b2c\u4e00\u4e2a\u975eTC\u5f00\u5934\u7684\u90e8\u5206
            for i in range(report_idx + 1, len(parts)):
                part = parts[i]
                # \u8df3\u8fc7TC\u5f00\u5934\u7684\u90e8\u5206\uff08\u5982TC4D9\uff09
                if part and not part.upper().startswith("TC"):
                    return part

        # \u515c\u5e95\uff1a\u5982\u679c\u6ca1\u627e\u5230\uff0c\u8fd4\u56de\u5012\u6570\u7b2c\u4e00\u4e2a\u975eTC\u5f00\u5934\u7684\u90e8\u5206
        for part in reversed(parts):
            if part and not part.upper().startswith("TC"):
                return part

        return ""

    def _find_header_map(self, rows: List[tuple], excel_columns: Dict) -> tuple[int, Dict[str, int]]:
        expected_id = self._normalize_header(excel_columns.get("id", "用例ID"))
        expected_trace = self._normalize_header(excel_columns.get("trace", "对应需求编号"))
        for idx, row in enumerate(rows[:20]):
            header_map: Dict[str, int] = {}
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                normalized = self._normalize_header(cell)
                if normalized:
                    header_map[normalized] = col_idx
            if not header_map:
                continue
            if expected_id in header_map and expected_trace in header_map:
                return idx, header_map
            # 兼容常见别名
            id_col = self._find_col(header_map, [excel_columns.get("id", ""), "用例ID", "测试用例ID", "用例编号", "ID"])
            trace_col = self._find_col(
                header_map,
                [excel_columns.get("trace", ""), "对应需求编号", "对应需求ID", "追溯需求", "关联需求", "需求编号", "需求ID"],
            )
            if id_col is not None and trace_col is not None:
                return idx, header_map
        return -1, {}

    def _find_col(self, header_map: Dict[str, int], candidates: List[str]) -> int | None:
        for candidate in candidates:
            normalized = self._normalize_header(candidate)
            if normalized and normalized in header_map:
                return header_map[normalized]
        return None

    @staticmethod
    def _safe_cell(row: tuple, col_idx: int | None) -> str:
        if col_idx is None or col_idx >= len(row):
            return ""
        value = row[col_idx]
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def extract_traced_keys(trace_cell_value: str) -> Set[str]:
        if not trace_cell_value:
            return set()
        normalized = str(trace_cell_value).upper()
        normalized = normalized.replace("－", "-").replace("—", "-").replace("–", "-")
        normalized = normalized.replace("_", "-")
        normalized = re.sub(r"\s*-\s*", "-", normalized)
        keys = ISSUE_KEY_PATTERN.findall(normalized)
        return set(keys)

    @staticmethod
    def _map_result(result_raw: str, mapping: Dict) -> str:
        value = str(result_raw or "").strip()
        if not value:
            return "not_executed"
        for target in ("passed", "failed", "not_executed"):
            candidates = {str(v).strip() for v in mapping.get(target, [])}
            if value in candidates:
                return target
        normalized = value.lower()
        if normalized in {"pass", "passed", "ok"}:
            return "passed"
        if normalized in {"fail", "failed", "ng"}:
            return "failed"
        return "not_executed"

    def dump_testcase_cache(self, testcases: List[TestCase]) -> Path:
        cache_file = self.cache_dir / "testcases" / "testcases_snapshot.json"
        cache_file.parent.mkdir(parents=True, exist_ok=True)
        payload = [
            {
                "case_id": tc.case_id,
                "case_name": tc.case_name,
                "test_type": tc.test_type,
                "traced_keys": sorted(tc.traced_keys),
                "result": tc.result,
                "source_file": tc.source_file,
                "source_page": tc.source_page,
                "module": tc.module,
            }
            for tc in testcases
        ]
        cache_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return cache_file
